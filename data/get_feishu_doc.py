#!/usr/bin/env python3
"""
飞书文档下载工具
================

从飞书下载文档或电子表格，支持自动查重并追加到主汇总 xlsx。

功能
----
  文档类型           下载格式                        额外操作
  docx（文档）       .txt → downloads/               无
  sheet（电子表格）  .csv/.json/.md → downloads/     查重后追加到主 xlsx
  wiki（知识库）     自动识别内部类型后按上方逻辑处理

快速开始
--------
  pip install requests openpyxl
  python get_feishu_doc.py

  支持链接格式：
    https://xxx.feishu.cn/wiki/Abc123   # Wiki 页面
    https://xxx.feishu.cn/docx/Abc123   # 文档
    Abc123                              # 直接填 token

目录结构
--------
  get-feishu-doc/
  ├── get_feishu_doc.py       本脚本
  ├── 外部待审核剧本.xlsx      主汇总表（查重追加目标）
  └── downloads/              所有下载文件
      ├── *.txt               docx 下载
      ├── *.csv               sheet 下载（UTF-8 BOM，Excel 可直接打开）
      ├── *.json              sheet 下载（结构化）
      └── *.md                sheet 下载（Markdown 表格）

查重逻辑
--------
  1. 读取 MASTER_XLSX 第 3 列（剧本名）的所有已有值
  2. 与飞书数据比对（忽略空格、全角空格）
  3. 仅将新行追加到 xlsx 末尾，重复行跳过
  4. 控制台输出：xlsx 更新: +N 新行，跳过 M 条重复

配置（脚本顶部常量）
--------------------
  FEISHU_APP_ID     飞书应用 App ID
  FEISHU_APP_SECRET 飞书应用 App Secret
  DOWNLOAD_DIR      下载目录（默认 downloads）
  MASTER_XLSX       主汇总表路径
  TITLE_COL_FEISHU  飞书数据中剧本名所在列（0-based，默认 2）
  TITLE_COL_XLSX    xlsx 中剧本名所在列（0-based，默认 2）

飞书应用所需权限
----------------
  wiki:wiki:readonly          读取知识库节点
  sheets:spreadsheet:readonly 读取电子表格
  docx:document:readonly      读取文档内容

依赖
----
  requests, openpyxl
"""

import re
import os
import json
import csv
import io
import time
import requests

# ── 配置 ──────────────────────────────────────────────────────────────────────

# 固定下载的飞书文档链接（多个表）
FEISHU_DOC_URLS = [
    "https://e76yjr9njh.feishu.cn/wiki/LNCEwl7gIiONSmkBa26cz1Gxnjf",  # 表1
    "https://e76yjr9njh.feishu.cn/wiki/Wqz7w7dvsiEjockJmqAcsnM2n4f",  # 表2
]

# 飞书应用凭证
FEISHU_APP_ID     = "cli_a955fe3f1e7a9bce"
FEISHU_APP_SECRET = "xwJC9KtmyvgQBfOi9mTKheLIfp7TG6O0"

# 文件路径
DOWNLOAD_DIR      = os.path.join(os.path.dirname(__file__), "downloads")
MASTER_XLSX       = os.path.join(os.path.dirname(__file__), "..", "uploads", "外部待审核剧本.xlsx")

# 查重用的列索引（0-based），对应"剧本"列
TITLE_COL_FEISHU  = 2   # Feishu 原始数据第 3 列
TITLE_COL_XLSX    = 2   # xlsx 第 3 列（0-based）

# 更新已存在行时不覆盖的列（0-based）。
# 评分列在 xlsx 里是 AVERAGE 公式（=AVERAGE(Gn:Xn)），飞书返回的是被求值后的文本，
# 且引用的行号是飞书侧的相对行号，直接覆盖会破坏公式。
PROTECTED_COLS    = {5}  # 第 6 列「评分（80+可签，70-80可改）」

# ─────────────────────────────────────────────────────────────────────────────

BASE_URL = "https://open.feishu.cn"
NO_PROXY = {"http": "", "https": ""}


# ── 认证 ─────────────────────────────────────────────────────────────────────

def get_tenant_access_token(app_id: str, app_secret: str) -> str:
    url = f"{BASE_URL}/open-apis/auth/v3/tenant_access_token/internal"
    resp = requests.post(url, json={"app_id": app_id, "app_secret": app_secret}, proxies=NO_PROXY)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        raise RuntimeError(f"获取 token 失败: {data.get('msg')}")
    return data["tenant_access_token"]


# ── URL / Token 解析 ──────────────────────────────────────────────────────────

def extract_doc_token(url_or_token: str) -> str:
    match = re.search(r"/(?:docx|wiki)/([A-Za-z0-9]+)", url_or_token)
    if match:
        return match.group(1)
    if "/" not in url_or_token:
        return url_or_token
    raise ValueError(f"无法从输入中解析文档 token: {url_or_token}")


def is_wiki_url(url_or_token: str) -> bool:
    return "/wiki/" in url_or_token


def resolve_wiki_node(token: str, wiki_token: str) -> tuple[str, str]:
    url = f"{BASE_URL}/open-apis/wiki/v2/spaces/get_node"
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers, params={"token": wiki_token}, proxies=NO_PROXY)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        raise RuntimeError(f"获取 Wiki 节点失败 (code={data.get('code')}): {data.get('msg')}")
    node = data["data"]["node"]
    return node["obj_token"], node["obj_type"]


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()


# ── docx 相关 ─────────────────────────────────────────────────────────────────

def get_doc_raw_content(token: str, doc_token: str) -> str:
    url = f"{BASE_URL}/open-apis/docx/v1/documents/{doc_token}/raw_content"
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers, proxies=NO_PROXY)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        raise RuntimeError(f"获取文档内容失败 (code={data.get('code')}): {data.get('msg')}")
    return data["data"]["content"]


def get_doc_title(token: str, doc_token: str) -> str:
    url = f"{BASE_URL}/open-apis/docx/v1/documents/{doc_token}"
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers, proxies=NO_PROXY)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        return doc_token
    return data["data"]["document"].get("title", doc_token)


# ── Sheet 相关 ────────────────────────────────────────────────────────────────

def get_sheet_meta(token: str, spreadsheet_token: str) -> dict:
    url = f"{BASE_URL}/open-apis/sheets/v3/spreadsheets/{spreadsheet_token}"
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers, proxies=NO_PROXY)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        raise RuntimeError(f"获取表格元信息失败 (code={data.get('code')}): {data.get('msg')}")
    return data["data"]["spreadsheet"]


def get_sheets_list(token: str, spreadsheet_token: str) -> list:
    url = f"{BASE_URL}/open-apis/sheets/v3/spreadsheets/{spreadsheet_token}/sheets/query"
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers, proxies=NO_PROXY)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        raise RuntimeError(f"获取子表列表失败 (code={data.get('code')}): {data.get('msg')}")
    return data["data"]["sheets"]


def get_sheet_values(token: str, spreadsheet_token: str, sheet_id: str) -> list:
    url = f"{BASE_URL}/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values/{sheet_id}"
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers, proxies=NO_PROXY)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        raise RuntimeError(f"获取表格数据失败 (code={data.get('code')}): {data.get('msg')}")
    return data["data"]["valueRange"].get("values", [])


# ── 单元格解析 ────────────────────────────────────────────────────────────────

def extract_cell_text(cell) -> str:
    """
    飞书 sheet API 返回的单元格可能是：
      - 字符串 / 数字
      - list of dicts（mention / text 片段）
    提取为纯文本字符串。
    """
    if cell is None:
        return ""
    if isinstance(cell, (int, float)):
        return str(cell)
    if isinstance(cell, str):
        return cell.strip()
    if isinstance(cell, list):
        parts = []
        for item in cell:
            if isinstance(item, dict):
                parts.append(str(item.get("text", "")).strip())
        return "".join(parts).strip()
    return str(cell).strip()


def normalize_title(title: str) -> str:
    """去除空白、全角空格，用于查重比较"""
    return re.sub(r"[\s\u3000]+", "", title)


# ── 格式转换 ──────────────────────────────────────────────────────────────────

def sheet_to_csv(rows: list) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow([extract_cell_text(cell) for cell in row])
    return buf.getvalue()


def rows_to_json(title: str, sheet_title: str, rows: list) -> dict:
    if not rows:
        return {"title": title, "sheet": sheet_title, "headers": [], "rows": []}
    headers = [extract_cell_text(c) for c in rows[0]]
    records = []
    for row in rows[1:]:
        record = {h: extract_cell_text(row[i] if i < len(row) else "") for i, h in enumerate(headers)}
        records.append(record)
    return {"title": title, "sheet": sheet_title, "headers": headers, "rows": records}


def rows_to_markdown(title: str, sheet_title: str, rows: list) -> str:
    lines = [f"# {title}", "", f"## {sheet_title}", ""]
    if not rows:
        return "\n".join(lines)
    headers = [extract_cell_text(c) for c in rows[0]]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows[1:]:
        cells = [extract_cell_text(row[i] if i < len(row) else "").replace("|", "\\|") for i in range(len(headers))]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


# ── xlsx 查重追加 ─────────────────────────────────────────────────────────────

def _load_workbook_with_title_map(xlsx_path: str):
    """打开 xlsx 并构建 {normalized_title: row_index(1-based)} 映射。
    文件不存在则返回 (None, None, {})，由调用方负责新建。
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("请先安装 openpyxl：pip install openpyxl")

    if not os.path.exists(xlsx_path):
        return None, None, {}

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    title_to_row = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        val = row[TITLE_COL_XLSX] if len(row) > TITLE_COL_XLSX else None
        if val:
            key = normalize_title(str(val))
            if key not in title_to_row:  # 遇到重复保留先出现的
                title_to_row[key] = idx
    return wb, ws, title_to_row


def sync_sheet_to_xlsx(rows: list, xlsx_path: str):
    """合并飞书数据到主 xlsx（飞书为准，空单元格保留 xlsx 原值）。

    rows[0] 为表头，rows[1:] 为数据行。策略：
      - 飞书行标题为空           → skipped（不处理）
      - 标题已存在 xlsx          → 按列覆盖，飞书该列为空则保留 xlsx 原值 → updated
      - 标题不存在 xlsx          → 整行追加 → added

    返回 (added, updated, skipped)。
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("请先安装 openpyxl：pip install openpyxl")

    wb, ws, title_to_row = _load_workbook_with_title_map(xlsx_path)
    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active

    added = updated = skipped = 0

    for row in rows[1:]:  # 跳过表头
        title_cell = row[TITLE_COL_FEISHU] if len(row) > TITLE_COL_FEISHU else None
        title = normalize_title(extract_cell_text(title_cell))
        if not title:
            skipped += 1
            continue

        cell_values = [extract_cell_text(cell) for cell in row]

        if title in title_to_row:
            row_idx = title_to_row[title]
            for col_idx_zero, new_val in enumerate(cell_values):
                # 空单元格保留 xlsx 原值；受保护列永不覆盖
                if new_val == "" or col_idx_zero in PROTECTED_COLS:
                    continue
                ws.cell(row=row_idx, column=col_idx_zero + 1).value = new_val
            updated += 1
        else:
            ws.append(cell_values)
            new_row_idx = ws.max_row
            # 新行的评分列用正确公式替换飞书的文本值
            if 5 in PROTECTED_COLS:
                ws.cell(row=new_row_idx, column=6).value = f"=AVERAGE(G{new_row_idx}:X{new_row_idx})"
            title_to_row[title] = new_row_idx  # 防本批次重复标题再次追加
            added += 1

    wb.save(xlsx_path)
    return added, updated, skipped


# ── 主下载逻辑 ────────────────────────────────────────────────────────────────

def step(n: int, total: int, msg: str, log=print):
    log(f"[{n}/{total}] {msg}")


def download_sheet(access_token: str, spreadsheet_token: str, output_dir: str = DOWNLOAD_DIR, log=print):
    """下载电子表格：CSV/JSON/MD 到 output_dir，并查重追加到主 xlsx"""
    os.makedirs(output_dir, exist_ok=True)

    log("  --> 获取表格元信息...", end=" ", flush=True)
    t0 = time.time()
    meta = get_sheet_meta(access_token, spreadsheet_token)
    title = sanitize_filename(meta.get("title", spreadsheet_token))
    log(f"完成 ({time.time()-t0:.1f}s)")

    log("  --> 获取子表列表...", end=" ", flush=True)
    t0 = time.time()
    sheets = get_sheets_list(access_token, spreadsheet_token)
    log(f"完成 ({time.time()-t0:.1f}s)  共 {len(sheets)} 个子表")

    all_saved = []
    for idx, sheet in enumerate(sheets, 1):
        sheet_id = sheet["sheet_id"]
        sheet_title = sanitize_filename(sheet.get("title", sheet_id))
        log(f"  --> [{idx}/{len(sheets)}] 拉取子表「{sheet_title}」数据...", end=" ", flush=True)
        t0 = time.time()
        rows = get_sheet_values(access_token, spreadsheet_token, sheet_id)
        log(f"完成 ({time.time()-t0:.1f}s)  {len(rows)} 行 x {max(len(r) for r in rows) if rows else 0} 列")

        base = f"{title}__{sheet_title}"

        log("  --> 写入 CSV / JSON / MD...", end=" ", flush=True)
        t0 = time.time()

        csv_path = os.path.join(output_dir, f"{base}.csv")
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write(sheet_to_csv(rows))

        json_path = os.path.join(output_dir, f"{base}.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(rows_to_json(title, sheet_title, rows), f, ensure_ascii=False, indent=2)

        md_path = os.path.join(output_dir, f"{base}.md")
        with open(md_path, "w", encoding="utf-8") as f:
            f.write(rows_to_markdown(title, sheet_title, rows))

        log(f"完成 ({time.time()-t0:.1f}s)")
        log(f"       {csv_path}")
        log(f"       {json_path}")
        log(f"       {md_path}")
        all_saved.extend([csv_path, json_path, md_path])

        log("  --> 查重并合并到主 xlsx...", end=" ", flush=True)
        t0 = time.time()
        added, updated, skipped = sync_sheet_to_xlsx(rows, MASTER_XLSX)
        log(f"完成 ({time.time()-t0:.1f}s)  +{added} 新行，更新 {updated} 条，跳过 {skipped} 条")

    return all_saved


def download_doc(url_or_token: str, output_path: str = None):
    total = 4
    t_start = time.time()

    step(1, total, "获取访问令牌...")
    t0 = time.time()
    access_token = get_tenant_access_token(FEISHU_APP_ID, FEISHU_APP_SECRET)
    print(f"     完成 ({time.time()-t0:.1f}s)")

    step(2, total, "解析文档 token...")
    t0 = time.time()
    raw_token = extract_doc_token(url_or_token)
    if is_wiki_url(url_or_token):
        print(f"     Wiki token: {raw_token}，解析实际文档节点...", end=" ", flush=True)
        doc_token, obj_type = resolve_wiki_node(access_token, raw_token)
        print(f"完成 ({time.time()-t0:.1f}s)  token={doc_token}  类型={obj_type}")
    else:
        doc_token = raw_token
        obj_type = "docx"
        print(f"     token: {doc_token}  类型: {obj_type}  ({time.time()-t0:.1f}s)")

    step(3, total, "获取文档内容...")
    if obj_type == "sheet":
        saved = download_sheet(access_token, doc_token)
        elapsed = time.time() - t_start
        print(f"\n完成！耗时 {elapsed:.1f}s，共保存 {len(saved)} 个文件:")
        for f in saved:
            print(f"  {f}")
        return

    t0 = time.time()
    content = get_doc_raw_content(access_token, doc_token)
    print(f"     完成 ({time.time()-t0:.1f}s)  {len(content)} 字符")

    if not output_path:
        title = get_doc_title(access_token, doc_token)
        filename = sanitize_filename(title) or doc_token
        os.makedirs(DOWNLOAD_DIR, exist_ok=True)
        output_path = os.path.join(DOWNLOAD_DIR, f"{filename}.txt")

    step(4, total, f"保存到 {output_path} ...")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    elapsed = time.time() - t_start
    print(f"\n完成！耗时 {elapsed:.1f}s  →  {output_path}")


# ── 入口 ──────────────────────────────────────────────────────────────────────

def run_sync(log_callback=None):
    """
    执行飞书文档拉取任务，返回执行结果字典。
    供定时任务或 API 手动触发调用。
    支持多个飞书文档 URL（依次下载）。

    返回:
        {
            "success": bool,
            "message": str,
            "files": list,
            "xlsx_added": int,
            "xlsx_skipped": int,
            "elapsed": float,
            "error": str or None
        }
    """
    result = {
        "success": False,
        "message": "",
        "files": [],
        "xlsx_added": 0,
        "xlsx_skipped": 0,
        "elapsed": 0,
        "error": None,
    }

    def log(msg="", **kwargs):
        print(msg, **kwargs)
        if log_callback:
            log_callback(msg)

    t_start = time.time()

    try:
        log(f"=== 飞书文档同步开始 ===")
        log(f"文档链接数: {len(FEISHU_DOC_URLS)}")
        for i, url in enumerate(FEISHU_DOC_URLS, 1):
            log(f"  [{i}] {url}")
        log(f"下载目录: {DOWNLOAD_DIR}/")
        log(f"主汇总表: {MASTER_XLSX}")
        log()

        step(1, 4, "获取访问令牌...", log)
        t0 = time.time()
        access_token = get_tenant_access_token(FEISHU_APP_ID, FEISHU_APP_SECRET)
        log(f"     完成 ({time.time()-t0:.1f}s)")

        total_files = []
        for url_idx, feishu_url in enumerate(FEISHU_DOC_URLS, 1):
            log(f"\n--- 处理文档 [{url_idx}/{len(FEISHU_DOC_URLS)}] ---")

            step(2, 4, "解析文档 token...", log)
            t0 = time.time()
            raw_token = extract_doc_token(feishu_url)
            if is_wiki_url(feishu_url):
                log(f"     Wiki token: {raw_token}，解析实际文档节点...")
                doc_token, obj_type = resolve_wiki_node(access_token, raw_token)
                log(f"完成 ({time.time()-t0:.1f}s)  token={doc_token}  类型={obj_type}")
            else:
                doc_token = raw_token
                obj_type = "docx"
                log(f"     token: {doc_token}  类型: {obj_type}  ({time.time()-t0:.1f}s)")

            if obj_type == "sheet":
                step(3, 4, "下载电子表格并查重...", log)
                saved_files = download_sheet(access_token, doc_token, log=log)
                total_files.extend(saved_files)
            else:
                step(3, 4, "获取文档内容...", log)
                t0 = time.time()
                content = get_doc_raw_content(access_token, doc_token)
                log(f"     完成 ({time.time()-t0:.1f}s)  {len(content)} 字符")

                title = get_doc_title(access_token, doc_token)
                filename = sanitize_filename(title) or doc_token
                os.makedirs(DOWNLOAD_DIR, exist_ok=True)
                output_path = os.path.join(DOWNLOAD_DIR, f"{filename}.txt")

                step(4, 4, f"保存到 {output_path} ...", log)
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(content)
                total_files.append(output_path)

        result["files"] = total_files
        result["xlsx_added"] = len(total_files)  # approximate
        result["elapsed"] = time.time() - t_start
        result["success"] = True
        result["message"] = f"同步完成，耗时 {result['elapsed']:.1f}s，共处理 {len(FEISHU_DOC_URLS)} 个文档"
        log(f"\n完成！耗时 {result['elapsed']:.1f}s，共保存 {len(total_files)} 个文件")

    except Exception as e:
        result["elapsed"] = time.time() - t_start
        result["error"] = str(e)
        result["message"] = f"同步失败: {e}"
        log(f"\n错误: {e}")

    return result


def main():
    run_sync()


if __name__ == "__main__":
    main()
